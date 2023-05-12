#coding=utf-8
import sys
import os
from   os.path import abspath, dirname
from turtle import left

from openpyxl.reader.excel import load_workbook

sys.path.append(abspath(dirname(__file__)))
import tkinter
import tkinter.filedialog
from   tkinter import *
import Fun
import numpy as np
import pandas as pd
ElementBGArray={}  
ElementBGArray_Resize={} 
ElementBGArray_IM={} 

def Pie_29_onLoadData(uiName,widgetName,Figure):
    a = Figure.add_subplot(111)

    # 读取excel表格
    wb = load_workbook('result.xlsx')
    sheet = wb.active
    qualified = 0
    unQualified = 0;
    for count in range(1, wb.worksheets[0].max_row+1):
        if sheet.cell(row=count, column=4).value == "Text content compliance Image review results: The images are not pornographic":
            qualified+=1
        else:
            unQualified+=1

    nums = [qualified,unQualified]
    labels = ['A','B']
    a.pie(x=nums,labels=labels)

    return wb.worksheets[0].max_row
def Bar_30_onLoadData(uiName,widgetName,Figure):
    a = Figure.add_subplot(111)
    a.bar(left=(0,1),height=(8,4),width=0.55)