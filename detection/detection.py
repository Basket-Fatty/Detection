# coding=utf-8
# import libs
import sys
import detection_cmd
import Fun
import os
os.environ["TF_CPP_MIN_LOG_LEVEL"]='1' # 这是默认的显示等级，显示所有信息
os.environ["TF_CPP_MIN_LOG_LEVEL"]='2' # 只显示 warning 和 Error
os.environ["TF_CPP_MIN_LOG_LEVEL"]='3' # 只显示 Error
import tkinter
from tkinter import *
import tkinter.ttk
import tkinter.font
from tkinter import filedialog
from PIL import Image
import re
import pandas as pd
from openpyxl import load_workbook
import xlwt
import xlrd
import datetime
from datetime import timedelta
from AC_automata import ac_automation
import numpy as np
from ocr_baidu import *



# Add your Varial Here: (Keep This Line of comments)
# Define UI Class
class Detection:
    def __init__(self, root, isTKroot=True):
        def openFile():
            '''打开选择文件夹对话框'''
            root1 = Tk()
            root1.withdraw()
            Filepath = filedialog.askopenfilename()  # 获得选择好的文件
            print(Filepath)
            Text_2.delete(1.0, 'end')  # 清除文本框内容
            Text_2.insert('insert', Filepath)  # 将结果添加到文本框显示
            # Folderpath = filedialog.askdirectory() #获得选择好的文件夹

        def OCR():

            n1 = Text_2.get(1.0, 'end')  # 获取文本框1的值

            result = n1[:-1]
            print(result)
            # 打开要识别的图片
            image = Image.open(result)
            # -----------------2/29新增内容--------------------------#
            # 调用ocr方法（百度文字识别API）识别文字
            stringList = final_ocr(result)
            text_save("text.txt", stringList)
            str_result = run("text.txt")
            if len(str_result)<27:
                str_result = '图片文字内容合格'
            result1 = str_result
            Text_3.insert('insert', result1+'\n' )  # 将结果添加到文本框显示


            # -------------------nsfw识别----------------------------#
            # 清除log.txt文件内容
            with open(r'log.txt', 'a+', encoding='utf-8') as log:
                log.truncate(0)

            picture_url = n1[:-1]
            p = ("python ./nsfw_master/classify_nsfw.py -m ./nsfw_master/data/open_nsfw-weights.npy ") + picture_url + (
                " >>log.txt")
            # 讲测试结果放入log.txt文件内
            os.system(p)

            # 读取log.txt
            with open(r'log.txt', mode='r', encoding='utf-8') as log:
                content1 = log.read()

            Text_3.insert('insert', content1)  # 将结果添加到文本框显示

            # 格式化时间
            time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            # 加载excel,注意路径要与脚本一致
            wb = load_workbook('result.xlsx')
            # 激活excel表
            sheet = wb.active

            count = wb.worksheets[0].max_row  # 总行数
            count += 1
            file = os.path.basename(picture_url)  # 输出图片名
            n = Text_3.get(1.0, 'end')  # 获取文本框1的值
            result = n[:-1]
            # 向excel中写入对应的value
            sheet.cell(row=count, column=1).value = count - 1
            sheet.cell(row=count, column=2).value = file
            sheet.cell(row=count, column=3).value = time
            sheet.cell(row=count, column=4).value = result

            wb.save('result.xlsx')
            print('数据写入成功！')

        def Batch_identification(file, path):
            # 打开要识别的图片
            # image = Image.open(path)
            # a = pytesseract.image_to_string(image, lang='chi_sim')
            stringlist =  final_ocr(path)
            text = ''
            text_save("text.txt", stringlist)
            str_result = run("text.txt")
            #text = "".join(stringlist.split())  # 去除字符串中所有的空格



            if len(str_result)<27 :
                str_result = '图片文字内容合格'

            result1 = str_result
            Text_3.insert('insert', result1+'\n')  # 将结果添加到文本框显示


            # -------------------nsfw识别----------------------------#
            # 清除log.txt文件内容
            with open(r'log.txt', 'a+', encoding='utf-8') as log:
                log.truncate(0)

            p = ("python ./nsfw_master/classify_nsfw.py -m ./nsfw_master/data/open_nsfw-weights.npy ") + path + (
                " >>log.txt")
            # 讲测试结果放入log.txt文件内
            os.system(p)

            # 读取log.txt
            with open(r'log.txt', mode='r') as log:
                content1 = log.read()

            '''
            str=str+','+content1
            data=[int(x) for x in str.split(',')]
            '''
            result1 = result1 + content1
            # 格式化时间
            time = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # -------------------------写入excel---------------------
            # 加载excel,注意路径要与脚本一致
            wb = load_workbook('result.xlsx')
            # 激活excel表
            sheet = wb.active

            count = wb.worksheets[0].max_row  # 总行数
            count += 1

            # 向excel中写入表头
            # sheet['a1'] = '姓名'
            # sheet['b1'] = '性别'
            # sheet['c1'] = '年龄'

            # 向excel中写入对应的value
            sheet.cell(row=count, column=1).value = count - 1
            sheet.cell(row=count, column=2).value = file
            sheet.cell(row=count, column=3).value = time
            sheet.cell(row=count, column=4).value = result1

            wb.save('result.xlsx')
            print('数据写入成功！')

        def findfiles():
            '''打开选择文件夹对话框'''
            root1 = Tk()
            root1.withdraw()
            # path = filedialog.askopenfilename() #获得选择好的文件
            path = filedialog.askdirectory()  # 获得选择好的文件夹
            Text_1.delete(1.0, 'end')  # 清除文本框内容
            Text_1.insert('insert', path)  # 将结果添加到文本框显示

        def identification():
            n1 = Text_1.get(1.0, 'end')  # 获取文本框1的值
            path = n1[:-1]
            # 首先遍历当前目录所有文件及文件夹
            file_list = os.listdir(path)
            # 循环判断每个元素是否是文件夹还是文件,是文件夹的话,递归
            for file in file_list:
                # 利用os.path.join()方法取得路径全名,并存入cur_path变量,否则每次只能遍历一层目录
                cur_path = os.path.join(path, file)
                print(file)

                Batch_identification(file, cur_path)

            Statistics(len(file_list))

        def openExcel():
            os.startfile('result.xlsx')

        def query():
            content = Entry_1.get()  # 获取文本框的值
            data = pd.read_excel('result.xlsx')
            result = data.loc[data['文件名'] == content]

            for i in range(len(result)):
                ListView_1.insert("", 1, text="line2", values=(
                int(result.iloc[i, [0]]['ID']), content, result.iloc[i, [2]]['更新时间'],
                result.iloc[i, [3]]['识别结果']))  # #给第0⾏添加数据,索引值可重复

        def Statistics(num):

            Canvas_1 = Fun.BuildChart('Pie', uiName, Form_1, 'Canvas_1')
            Fun.Register(uiName, 'Canvas_1', Canvas_1)
            Fun.SetControlPlace(uiName, 'Canvas_1', 20, 240, 240, 240)
            detection_cmd.Pie_29_onLoadData(uiName, 'Canvas_1', Fun.GetUserData(uiName, 'Canvas_1', 'ChartFigure'))

            Canvas_2 = Fun.BuildChart('Bar', uiName, Form_1, 'Canvas_2')
            Fun.Register(uiName, 'Canvas_2', Canvas_2)
            Fun.SetControlPlace(uiName, 'Canvas_2', 304, 240, 240, 240)
            detection_cmd.Bar_30_onLoadData(uiName, 'Canvas_2', Fun.GetUserData(uiName, 'Canvas_2', 'ChartFigure'))
            Label_9 = tkinter.Label(Form_1, text="Illegal")
            Fun.Register(uiName, 'Label_9', Label_9, 'Illegal Category')
            Fun.SetControlPlace(uiName, 'Label_9', 320, 455, 100, 30)
            Label_9.configure(bg="#ffffff")
            Label_9.configure(relief="flat")
            Label_9_Ft = tkinter.font.Font(family='华文新魏', size=12, weight='normal', slant='roman', underline=0,
                                           overstrike=0)
            Label_9.configure(font=Label_9_Ft)
            Label_10 = tkinter.Label(Form_1, text="Qualified")
            Fun.Register(uiName, 'Label_10', Label_10, 'Qualified Category')
            Fun.SetControlPlace(uiName, 'Label_10', 400, 455, 180, 30)
            Label_10.configure(bg="#ffffff")
            Label_10.configure(relief="flat")
            Label_10_Ft = tkinter.font.Font(family='华文新魏', size=12, weight='normal', slant='roman', underline=0,
                                            overstrike=0)
            Label_10.configure(font=Label_10_Ft)

            Label_12 = tkinter.Label(Form_1, text="Number" + str(num) + "pictures, results are following:")
            Fun.Register(uiName, 'Label_12', Label_12, 'Qualified Category')
            Fun.SetControlPlace(uiName, 'Label_12', 100, 200, 400, 30)
            Label_12.configure(bg="#ffffff")
            Label_12.configure(relief="flat")
            Label_12_Ft = tkinter.font.Font(family='华文新魏', size=12, weight='normal', slant='roman', underline=0,
                                            overstrike=0)
            Label_12.configure(font=Label_12_Ft)

            Label_13 = tkinter.Label(Form_1, text="A-Cosmetics,B-Health Care Products,C-Daily Necessities,\nD-Medicine,E-Education")
            Fun.Register(uiName, 'Label_13', Label_13, 'legend')
            Fun.SetControlPlace(uiName, 'Label_13', 15, 450, 240, 30)
            Label_13.configure(bg="#ffffff")
            Label_13.configure(relief="flat")
            Label_13_Ft = tkinter.font.Font(family='华文新魏', size=12, weight='normal', slant='roman', underline=0,
                                            overstrike=0)
            Label_13.configure(font=Label_13_Ft)

        uiName = self.__class__.__name__
        self.uiName = uiName
        Fun.Register(uiName, 'UIClass', self)
        self.root = root
        self.isTKroot = isTKroot
        Fun.Register(uiName, 'root', root)
        # style = detection_sty.SetupStyle()
        if isTKroot == True:
            root.title("INTERNET ILLEGAL ADVERTISEMENTS RECOGNITION")
            Fun.CenterDlg(uiName, root, 1200, 600)
            root['background'] = '#efefef'
        Form_1 = tkinter.Canvas(root, width=10, height=4)
        Form_1.pack(fill=BOTH, expand=True)
        Form_1.configure(bg="#efefef")
        Form_1.configure(highlightthickness=0)
        Fun.Register(uiName, 'Form_1', Form_1)
        # Create the elements of root
        Label_1 = tkinter.Label(Form_1, text="INTERNET ILLEGAL ADVERTISEMENTS RECOGNITION")
        Fun.Register(uiName, 'Label_1', Label_1, 'INTERNET ILLEGAL ADVERTISEMENTS RECOGNITION')
        Fun.SetControlPlace(uiName, 'Label_1', 250, 0, 737, 50)
        Label_1.configure(relief="flat")
        Label_1_Ft = tkinter.font.Font(family='华文新魏', size=20, weight='bold', slant='roman', underline=0, overstrike=0)
        Label_1.configure(font=Label_1_Ft)
        ListBox_1 = tkinter.Listbox(Form_1)
        Fun.Register(uiName, 'ListBox_1', ListBox_1, 'Single Recognition')
        Fun.SetControlPlace(uiName, 'ListBox_1', 600, 50, 600, 300)
        ListBox_2 = tkinter.Listbox(Form_1)
        Fun.Register(uiName, 'ListBox_2', ListBox_2, 'Batch Process')
        Fun.SetControlPlace(uiName, 'ListBox_2', 0, 50, 600, 550)
        Label_2 = tkinter.Label(Form_1, text="Batch Process")
        Fun.Register(uiName, 'Label_2', Label_2, 'Batch Process')
        Fun.SetControlPlace(uiName, 'Label_2', 0, 50, 100, 20)
        Label_2.configure(relief="flat")
        Label_3 = tkinter.Label(Form_1, text="Single Process")
        Fun.Register(uiName, 'Label_3', Label_3, 'Single Process')
        Fun.SetControlPlace(uiName, 'Label_3', 600, 50, 100, 20)
        Label_3.configure(relief="flat")
        Label_4 = tkinter.Label(Form_1, text="Folder Path:")
        Fun.Register(uiName, 'Label_4', Label_4, 'Folder Path')
        Fun.SetControlPlace(uiName, 'Label_4', 20, 80, 150, 40)
        Label_4.configure(bg="#ffffff")
        Label_4.configure(relief="flat")
        Label_4_Ft = tkinter.font.Font(family='华文新魏', size=12, weight='normal', slant='roman', underline=0,
                                       overstrike=0)
        Label_4.configure(font=Label_4_Ft)
        Text_1 = tkinter.Text(Form_1)
        Fun.Register(uiName, 'Text_1', Text_1, 'Enter Path')
        Fun.SetControlPlace(uiName, 'Text_1', 170, 80, 150, 40)
        Text_1.configure(bg="#ffffff")
        Text_1.configure(relief="sunken")
        Button_1 = tkinter.Button(Form_1, text="Choose", command=findfiles)
        Fun.Register(uiName, 'Button_1', Button_1, 'Choose')
        Fun.SetControlPlace(uiName, 'Button_1', 320, 80, 120, 40)
        Button_1_Ft = tkinter.font.Font(family='华文新魏', size=12, weight='normal', slant='roman', underline=0,
                                        overstrike=0)
        Button_1.configure(font=Button_1_Ft)
        Button_2 = tkinter.Button(Form_1, text="Recognition", command=identification)
        Fun.Register(uiName, 'Button_2', Button_2, 'Recognition')
        Fun.SetControlPlace(uiName, 'Button_2', 440, 80, 100, 40)
        Button_2_Ft = tkinter.font.Font(family='华文新魏', size=12, weight='normal', slant='roman', underline=0,
                                        overstrike=0)
        Button_2.configure(font=Button_2_Ft)
        Label_5 = tkinter.Label(Form_1, text="Picture Path:")
        Fun.Register(uiName, 'Label_5', Label_5, 'Picture Path')
        Fun.SetControlPlace(uiName, 'Label_5', 620, 80, 150, 40)
        Label_5.configure(bg="#ffffff")
        Label_5.configure(relief="flat")
        Label_5_Ft = tkinter.font.Font(family='华文新魏', size=12, weight='normal', slant='roman', underline=0,
                                       overstrike=0)
        Label_5.configure(font=Label_5_Ft)
        ListBox_3 = tkinter.Listbox(Form_1)
        Fun.Register(uiName, 'ListBox_3', ListBox_3, 'History')
        Fun.SetControlPlace(uiName, 'ListBox_3', 600, 350, 600, 250)
        Button_3 = tkinter.Button(Form_1, text="Choose", command=openFile)
        Fun.Register(uiName, 'Button_3', Button_3, 'Choose')
        Fun.SetControlPlace(uiName, 'Button_3', 920, 80, 100, 40)
        Button_3_Ft = tkinter.font.Font(family='华文新魏', size=12, weight='normal', slant='roman', underline=0,
                                        overstrike=0)
        Button_3.configure(font=Button_3_Ft)
        Button_4 = tkinter.Button(Form_1, text="Recognition", command=OCR)
        Fun.Register(uiName, 'Button_4', Button_4, 'Single Recognition')
        Fun.SetControlPlace(uiName, 'Button_4', 1020, 80, 100, 40)
        Button_4_Ft = tkinter.font.Font(family='华文新魏', size=12, weight='normal', slant='roman', underline=0,
                                        overstrike=0)
        Button_4.configure(font=Button_4_Ft)
        Label_6 = tkinter.Label(Form_1, text="Result:")
        Fun.Register(uiName, 'Label_6', Label_6, 'Single Result')
        Fun.SetControlPlace(uiName, 'Label_6', 620, 120, 150, 40)
        Label_6.configure(bg="#ffffff")
        Label_6.configure(relief="flat")
        Label_6_Ft = tkinter.font.Font(family='华文新魏', size=12, weight='normal', slant='roman', underline=0,
                                       overstrike=0)
        Label_6.configure(font=Label_6_Ft)
        Text_2 = tkinter.Text(Form_1)
        Fun.Register(uiName, 'Text_2', Text_2, 'Access Image Address')
        Fun.SetControlPlace(uiName, 'Text_2', 770, 80, 150, 40)
        Text_2.configure(relief="sunken")
        Text_3 = tkinter.Text(Form_1)
        Fun.Register(uiName, 'Text_3', Text_3, 'Picture Result')
        Fun.SetControlPlace(uiName, 'Text_3', 770, 130, 350, 160)
        Text_3.configure(relief="sunken")
        Label_6 = tkinter.Label(Form_1, text="History")
        Fun.Register(uiName, 'Label_6', Label_6, 'History Query')
        Fun.SetControlPlace(uiName, 'Label_6', 600, 350, 100, 20)
        Label_5.configure(relief="flat")
        Label_7 = tkinter.Label(Form_1, text="Result:")
        Fun.Register(uiName, 'Label_7', Label_7, 'Batch Result')
        Fun.SetControlPlace(uiName, 'Label_7', 20, 120, 150, 40)
        Label_7.configure(bg="#ffffff")
        Label_7.configure(relief="flat")
        Label_7_Ft = tkinter.font.Font(family='华文新魏', size=12, weight='normal', slant='roman', underline=0,
                                       overstrike=0)
        Label_7.configure(font=Label_7_Ft)
        ListView_1 = tkinter.ttk.Treeview(Form_1, show="headings")
        Fun.Register(uiName, 'ListView_1', ListView_1)
        Fun.SetControlPlace(uiName, 'ListView_1', 620, 420, 560, 160)
        ListView_1.configure(selectmode="extended")
        ListView_1.configure(columns=["ID", "File Name", "Last Update Time", "Result"])
        ListView_1.column("ID", anchor="center", width=20)
        ListView_1.heading("ID", anchor="center", text="ID")
        ListView_1.column("File Name", anchor="center", width=60)
        ListView_1.heading("File Name", anchor="center", text="File Name")
        ListView_1.column("Last Update Time", anchor="center", width=60)
        ListView_1.heading("Last Update Time", anchor="center", text="Last Update Time")
        ListView_1.column("Result", anchor="center", width=150)
        ListView_1.heading("Result", anchor="center", text="Result")

        Label_8 = tkinter.Label(Form_1, text="Picture Name:")
        Fun.Register(uiName, 'Label_8', Label_8, 'File Name')
        Fun.SetControlPlace(uiName, 'Label_8', 652, 385, 100, 20)
        Label_8.configure(relief="flat")
        Entry_1_Variable = Fun.AddTKVariable(uiName, 'Entry_1', '')
        Entry_1 = tkinter.Entry(Form_1, textvariable=Entry_1_Variable)
        Fun.Register(uiName, 'Entry_1', Entry_1, 'Picture Name')
        Fun.SetControlPlace(uiName, 'Entry_1', 752, 385, 120, 20)
        Entry_1.configure(relief="sunken")
        Button_5 = tkinter.Button(Form_1, text="Search", command=query)
        Fun.Register(uiName, 'Button_5', Button_5, 'Search')
        Fun.SetControlPlace(uiName, 'Button_5', 865, 384, 100, 28)
        Button_6 = tkinter.Button(Form_1, text="More", command=openExcel)
        Fun.Register(uiName, 'Button_6', Button_6, 'More')
        Fun.SetControlPlace(uiName, 'Button_6', 990, 383, 100, 28)

        # Add Some Logic Code Here: (Keep This Line of comments)
        # Statistics(3)

        # Exit Application: (Keep This Line of comments)
        if self.isTKroot == True:
            self.root.protocol('WM_DELETE_WINDOW', self.exit)

    def exit(self):
        if self.isTKroot == True:
            self.root.destroy()


# Create the root of Kinter
if __name__ == '__main__':
    root = tkinter.Tk()
    MyDlg = Detection(root)
    root.mainloop()

