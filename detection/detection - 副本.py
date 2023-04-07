#coding=utf-8
#import libs 
import sys
import detection_cmd
import detection_sty
import Fun
import os
import tkinter
from   tkinter import *
import tkinter.ttk
import tkinter.font
from tkinter import filedialog
from PIL import Image
import re
import pandas as pd
from openpyxl import load_workbook
import xlwt
import xlrd
import datetime
from datetime import timedelta

#Add your Varial Here: (Keep This Line of comments)
#Define UI Class
class  detection:
    def __init__(self,root,isTKroot = True):
        def openFile():
            '''打开选择文件夹对话框'''
            root1 = Tk()
            root1.withdraw()
            Filepath = filedialog.askopenfilename() #获得选择好的文件
            Text_2.delete(1.0, 'end')  # 清除文本框内容
            Text_2.insert('insert', Filepath)  # 将结果添加到文本框显示
            #Folderpath = filedialog.askdirectory() #获得选择好的文件夹
        
        def OCR():
            # 导入OCR安装路径，如果设置了系统环境，就可以不用设置了
            #pytesseract.pytesseract.tesseract_cmd = r"D:\Tesseract-OCR\tesseract.exe"
            n1 = Text_2.get(1.0, 'end')  # 获取文本框1的值
            result=n1[:-1]
            # 打开要识别的图片
            image = Image.open(result)
            a = pytesseract.image_to_string(image,lang='chi_sim')
            #b = a.split() # 字符串按空格分割成列表
            #c = "".join(b) # 使用一个空字符串合成列表内容生成新的字符串
            text="".join(a.split())#去除字符串中所有的空格
            #print(text)

            fname ="./error.txt"
            with open(fname, mode='r',encoding='gbk') as f: # 打开文件
                lines = f.readlines() # 读取所有行
                #print(lines)
                l=len(lines)
                #print(l)
                for i in range(0,l):
                    pattern=lines[i][:-1]#读取第l行
                    pattern = re.findall('[\u4e00-\u9fa5]', pattern)
                    pattern="".join(pattern)
                    #print(pattern)
                    m=re.search(pattern,text)
                    #print(m)
                    if m==None :
                        i+=1
                    else:
                        result="广告内有违规词："+lines[i]
                        #print(result)
                        Text_3.insert('insert', result)  # 将结果添加到文本框显示
                        break
            
            #-------------------nsfw识别----------------------------#
            #清除log.txt文件内容
            with open(r'log.txt','a+',encoding='utf-8') as log:
                log.truncate(0)

            picture_url=n1[:-1]
            p=("python ./nsfw_master/classify_nsfw.py -m ./nsfw_master/data/open_nsfw-weights.npy ")+picture_url+(" >>log.txt")
            #讲测试结果放入log.txt文件内
            os.system(p)
    
            #读取log.txt
            with open(r'log.txt',mode='r') as log:
                content1=log.read()

            Text_3.insert('insert', content1)  # 将结果添加到文本框显示


            '''
            n2 = Text_2.get(1.0, 'end')  # 获取文本框1的值
            txt=n2[:-4]+'txt'
            #读取log.txt
            with open(txt,mode='w') as file:
                file.write(text)
        
            result2=n2[:-4]+'txt'
            t2.delete(1.0, 'end')  # 清除文本框内容
            t2.insert('insert', result2)  # 将结果添加到文本框显示
            '''
        def Batch_identification(file,path):
            # 打开要识别的图片
            image = Image.open(path)
            a = pytesseract.image_to_string(image,lang='chi_sim')
            text="".join(a.split())#去除字符串中所有的空格

            fname ="./error.txt"
            with open(fname, mode='r',encoding='gbk') as f: # 打开文件
                lines = f.readlines() # 读取所有行
                l=len(lines)
                for i in range(0,l):
                    pattern=lines[i][:-1]#读取第l行
                    pattern = re.findall('[\u4e00-\u9fa5]', pattern)
                    pattern="".join(pattern)
                    m=re.search(pattern,text)
                    if m==None :
                        i+=1
                    else:
                        result="广告内有违规词："+lines[i]
                        #print(result)
                        #Text_3.insert('insert', result)  # 将结果添加到文本框显示
                        
                        break
            
            #-------------------nsfw识别----------------------------#
            #清除log.txt文件内容
            with open(r'log.txt','a+',encoding='utf-8') as log:
                log.truncate(0)

            p=("python ./nsfw_master/classify_nsfw.py -m ./nsfw_master/data/open_nsfw-weights.npy ")+path+(" >>log.txt")
            #讲测试结果放入log.txt文件内
            os.system(p)
    
            #读取log.txt
            with open(r'log.txt',mode='r') as log:
                content1=log.read()
            
            '''
            str=str+','+content1
            data=[int(x) for x in str.split(',')]
            '''
            result=result+content1
            # 格式化时间
            time=datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            #-------------------------写入excel---------------------
            #加载excel，注意路径要与脚本一致
            wb = load_workbook('result.xlsx')
            #激活excel表
            sheet = wb.active

            count=wb.worksheets[0].max_row#总行数
            count+=1
 
            #向excel中写入表头
            #sheet['a1'] = '姓名'
            #sheet['b1'] = '性别'
            #sheet['c1'] = '年龄'
 
            #向excel中写入对应的value
            sheet.cell(row=count, column=1).value = count-1
            sheet.cell(row=count, column=2).value = file
            sheet.cell(row=count, column=3).value = time
            sheet.cell(row=count, column=4).value = result
 
            wb.save('result.xlsx')
            print('数据写入成功！')

        def findfiles():
            '''打开选择文件夹对话框'''
            root1 = Tk()
            root1.withdraw()
            #path = filedialog.askopenfilename() #获得选择好的文件
            path = filedialog.askdirectory() #获得选择好的文件夹
            Text_1.delete(1.0, 'end')  # 清除文本框内容
            Text_1.insert('insert', path)  # 将结果添加到文本框显示

        def identification():
            n1 = Text_1.get(1.0, 'end')  # 获取文本框1的值
            path=n1[:-1]
            # 首先遍历当前目录所有文件及文件夹
            file_list = os.listdir(path)
            # 循环判断每个元素是否是文件夹还是文件，是文件夹的话，递归
            for file in file_list:
    	        # 利用os.path.join()方法取得路径全名，并存入cur_path变量，否则每次只能遍历一层目录
                cur_path = os.path.join(path, file)
                print(file)
                Batch_identification(file,cur_path)
                
        def openExcel():
            os.startfile('result.xlsx')

        def query():
            content = Entry_1.get()  # 获取文本框的值
            data=pd.read_excel('result.xlsx')
            result=data.loc[data['文件名']==content]
            
            for i in range(len(result)):
                ListView_1.insert("", 1, text="line2", values=(int(result.iloc[i, [0]]['ID']),content,result.iloc[i, [2]]['更新时间'],result.iloc[i, [3]]['识别结果']))    # #给第0⾏添加数据，索引值可重复
            

        uiName = self.__class__.__name__
        self.uiName = uiName
        Fun.Register(uiName,'UIClass',self)
        self.root = root
        self.isTKroot = isTKroot
        Fun.Register(uiName,'root',root)
        style = detection_sty.SetupStyle()
        if isTKroot == True:
            root.title("融媒体违规广告识别软件")
            Fun.CenterDlg(uiName,root,1200,600)
            root['background'] = '#efefef'
        Form_1= tkinter.Canvas(root,width = 10,height = 4)
        Form_1.pack(fill=BOTH,expand=True)
        Form_1.configure(bg = "#efefef")
        Form_1.configure(highlightthickness = 0)
        Fun.Register(uiName,'Form_1',Form_1)
        #Create the elements of root 
        Label_1 = tkinter.Label(Form_1,text="融媒体违规广告识别软件")
        Fun.Register(uiName,'Label_1',Label_1,'融媒体违规广告识别软件')
        Fun.SetControlPlace(uiName,'Label_1',250,0,737,50)
        Label_1.configure(relief = "flat")
        Label_1_Ft=tkinter.font.Font(family='华文新魏', size=25,weight='bold',slant='roman',underline=0,overstrike=0)
        Label_1.configure(font = Label_1_Ft)
        ListBox_1 = tkinter.Listbox(Form_1)
        Fun.Register(uiName,'ListBox_1',ListBox_1,'单张识别')
        Fun.SetControlPlace(uiName,'ListBox_1',600,50,600,300)
        ListBox_2 = tkinter.Listbox(Form_1)
        Fun.Register(uiName,'ListBox_2',ListBox_2,'批次处理')
        Fun.SetControlPlace(uiName,'ListBox_2',0,50,600,550)
        Label_2 = tkinter.Label(Form_1,text="批处理")
        Fun.Register(uiName,'Label_2',Label_2,'批处理')
        Fun.SetControlPlace(uiName,'Label_2',0,50,100,20)
        Label_2.configure(relief = "flat")
        Label_3 = tkinter.Label(Form_1,text="单一处理")
        Fun.Register(uiName,'Label_3',Label_3,'单一处理')
        Fun.SetControlPlace(uiName,'Label_3',600,50,100,20)
        Label_3.configure(relief = "flat")
        Label_4 = tkinter.Label(Form_1,text="文件夹路径：")
        Fun.Register(uiName,'Label_4',Label_4,'文件夹路径')
        Fun.SetControlPlace(uiName,'Label_4',20,80,150,40)
        Label_4.configure(bg = "#ffffff")
        Label_4.configure(relief = "flat")
        Label_4_Ft=tkinter.font.Font(family='华文新魏', size=15,weight='normal',slant='roman',underline=0,overstrike=0)
        Label_4.configure(font = Label_4_Ft)
        Text_1 = tkinter.Text(Form_1)
        Fun.Register(uiName,'Text_1',Text_1,'输入地址')
        Fun.SetControlPlace(uiName,'Text_1',170,80,150,40)
        Text_1.configure(bg = "#ffffff")
        Text_1.configure(relief = "sunken")
        Button_1 = tkinter.Button(Form_1,text="选择文件",command=findfiles)
        Fun.Register(uiName,'Button_1',Button_1,'选择文件')
        Fun.SetControlPlace(uiName,'Button_1',320,80,120,40)
        Button_1_Ft=tkinter.font.Font(family='华文新魏', size=13,weight='normal',slant='roman',underline=0,overstrike=0)
        Button_1.configure(font = Button_1_Ft)
        Button_2 = tkinter.Button(Form_1,text="识别",command=identification)
        Fun.Register(uiName,'Button_2',Button_2,'识别')
        Fun.SetControlPlace(uiName,'Button_2',440,80,100,40)
        Button_2_Ft=tkinter.font.Font(family='华文新魏', size=13,weight='normal',slant='roman',underline=0,overstrike=0)
        Button_2.configure(font = Button_2_Ft)
        Label_5 = tkinter.Label(Form_1,text="图片路径：")
        Fun.Register(uiName,'Label_5',Label_5,'图片路径')
        Fun.SetControlPlace(uiName,'Label_5',620,80,150,40)
        Label_5.configure(bg = "#ffffff")
        Label_5.configure(relief = "flat")
        Label_5_Ft=tkinter.font.Font(family='华文新魏', size=15,weight='normal',slant='roman',underline=0,overstrike=0)
        Label_5.configure(font = Label_5_Ft)
        ListBox_3 = tkinter.Listbox(Form_1)
        Fun.Register(uiName,'ListBox_3',ListBox_3,'历史记录')
        Fun.SetControlPlace(uiName,'ListBox_3',600,350,600,250)
        Button_3 = tkinter.Button(Form_1,text="选择图片",command=openFile)
        Fun.Register(uiName,'Button_3',Button_3,'图片选择')
        Fun.SetControlPlace(uiName,'Button_3',920,80,100,40)
        Button_3_Ft=tkinter.font.Font(family='华文新魏', size=13,weight='normal',slant='roman',underline=0,overstrike=0)
        Button_3.configure(font = Button_3_Ft)
        Button_4 = tkinter.Button(Form_1,text="识别",command=OCR)
        Fun.Register(uiName,'Button_4',Button_4,'单一识别')
        Fun.SetControlPlace(uiName,'Button_4',1020,80,100,40)
        Button_4_Ft=tkinter.font.Font(family='华文新魏', size=13,weight='normal',slant='roman',underline=0,overstrike=0)
        Button_4.configure(font = Button_4_Ft)
        Label_6 = tkinter.Label(Form_1,text="识别结果：")
        Fun.Register(uiName,'Label_6',Label_6,'单一结果')
        Fun.SetControlPlace(uiName,'Label_6',620,120,150,40)
        Label_6.configure(bg = "#ffffff")
        Label_6.configure(relief = "flat")
        Label_6_Ft=tkinter.font.Font(family='华文新魏', size=15,weight='normal',slant='roman',underline=0,overstrike=0)
        Label_6.configure(font = Label_6_Ft)
        Text_2 = tkinter.Text(Form_1)
        Fun.Register(uiName,'Text_2',Text_2,'访问图片地址')
        Fun.SetControlPlace(uiName,'Text_2',770,80,150,40)
        Text_2.configure(relief = "sunken")
        Text_3 = tkinter.Text(Form_1)
        Fun.Register(uiName,'Text_3',Text_3,'图片识别结果')
        Fun.SetControlPlace(uiName,'Text_3',770,130,350,160)
        Text_3.configure(relief = "sunken")
        Label_6 = tkinter.Label(Form_1,text="历史记录")
        Fun.Register(uiName,'Label_6',Label_6,'历史记录查询')
        Fun.SetControlPlace(uiName,'Label_6',600,350,100,20)
        Label_5.configure(relief = "flat")
        Label_7 = tkinter.Label(Form_1,text="识别结果：")
        Fun.Register(uiName,'Label_7',Label_7,'批量结果')
        Fun.SetControlPlace(uiName,'Label_7',20,120,150,40)
        Label_7.configure(bg = "#ffffff")
        Label_7.configure(relief = "flat")
        Label_7_Ft=tkinter.font.Font(family='华文新魏', size=15,weight='normal',slant='roman',underline=0,overstrike=0)
        Label_7.configure(font = Label_7_Ft)
        ListView_1 = tkinter.ttk.Treeview(Form_1,show="headings")
        Fun.Register(uiName,'ListView_1',ListView_1)
        Fun.SetControlPlace(uiName,'ListView_1',620,420,560,160)
        ListView_1.configure(selectmode = "extended")
        ListView_1.configure(columns = ["ID","文件名","更新时间","识别结果"])
        ListView_1.column("ID",anchor="center",width=20)
        ListView_1.heading("ID",anchor="center",text="ID")
        ListView_1.column("文件名",anchor="center",width=60)
        ListView_1.heading("文件名",anchor="center",text="文件名")
        ListView_1.column("更新时间",anchor="center",width=60)
        ListView_1.heading("更新时间",anchor="center",text="更新时间")
        ListView_1.column("识别结果",anchor="center",width=150)
        ListView_1.heading("识别结果",anchor="center",text="识别结果")
        
        #ListView_1.bind("<<TreeviewSelect>>",Fun.EventFunction_Adaptor(detection_cmd.ListView_1_onSelect,uiName=uiName,widgetName="ListView_1"))
        Label_8 = tkinter.Label(Form_1,text="图片名：")
        Fun.Register(uiName,'Label_8',Label_8,'文件名')
        Fun.SetControlPlace(uiName,'Label_8',652,385,100,20)
        Label_8.configure(relief = "flat")
        Entry_1_Variable = Fun.AddTKVariable(uiName,'Entry_1','')
        Entry_1 = tkinter.Entry(Form_1,textvariable=Entry_1_Variable)
        Fun.Register(uiName,'Entry_1',Entry_1,'查询图片名')
        Fun.SetControlPlace(uiName,'Entry_1',752,385,120,20)
        Entry_1.configure(relief = "sunken")
        Button_5 = tkinter.Button(Form_1,text="查询",command=query)
        Fun.Register(uiName,'Button_5',Button_5,'查询')
        Fun.SetControlPlace(uiName,'Button_5',865,384,100,28)
        Button_6 = tkinter.Button(Form_1,text="更多",command=openExcel)
        Fun.Register(uiName,'Button_6',Button_6,'更多')
        Fun.SetControlPlace(uiName,'Button_6',990,383,100,28)
        Canvas_1 = Fun.BuildChart('Pie',uiName,Form_1,'Canvas_1')
        Fun.Register(uiName,'Canvas_1',Canvas_1)
        Fun.SetControlPlace(uiName,'Canvas_1',20,300,240,240)
        detection_cmd.Pie_29_onLoadData(uiName,'Canvas_1',Fun.GetUserData(uiName,'Canvas_1','ChartFigure'))
        Canvas_2 = Fun.BuildChart('Bar',uiName,Form_1,'Canvas_2')
        Fun.Register(uiName,'Canvas_2',Canvas_2)
        Fun.SetControlPlace(uiName,'Canvas_2',304,300,240,240)
        detection_cmd.Bar_30_onLoadData(uiName,'Canvas_2',Fun.GetUserData(uiName,'Canvas_2','ChartFigure'))
        #Create the Menu of root 
        '''
        MainMenu=tkinter.Menu(root)
        root.config(menu = MainMenu)
        文件名=tkinter.Menu(MainMenu,tearoff = 0)
        文件名.add_command(label="文件名1",command=lambda:detection_cmd.Menu_文件名1(uiName,"文件名1"))
        MainMenu.add_cascade(label="文件名",menu=文件名)
        #Inital all element's Data 
        Fun.InitElementData(uiName)
        '''
        #Add Some Logic Code Here: (Keep This Line of comments)


        #Exit Application: (Keep This Line of comments)
        if self.isTKroot == True:
            self.root.protocol('WM_DELETE_WINDOW', self.exit)

    def exit(self):
        if self.isTKroot == True:
            self.root.destroy()

#Create the root of Kinter 
if  __name__ == '__main__':
    root = tkinter.Tk()
    MyDlg = detection(root)
    root.mainloop()
